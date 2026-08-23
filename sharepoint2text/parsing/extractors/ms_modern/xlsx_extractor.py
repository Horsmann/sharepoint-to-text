"""
XLSX/XLSB Spreadsheet Extractor

Extracts text content and metadata from Microsoft Excel spreadsheet files.

- .xlsx/.xlsm: Parsed with openpyxl.
- .xlsb: Parsed with pyxlsb.
"""

import datetime
import io
import logging
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Generator, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]
from pyxlsb import open_workbook  # type: ignore[import-untyped]

from sharepoint2text.parsing.exceptions import (
    ExtractionError,
    ExtractionFailedError,
    ExtractionFileEncryptedError,
)
from sharepoint2text.parsing.extractors._model import source_metadata
from sharepoint2text.parsing.extractors.ms_modern.ooxml_namespaces import (
    A_BLIP_XLSX,
    ANCHOR_TYPES,
    EMU_PER_PIXEL,
    R_EMBED_XLSX,
    XDR_BLIPFILL,
    XDR_CNVPR,
    XDR_EXT,
    XDR_NVPICPR,
    XDR_PIC,
)
from sharepoint2text.parsing.extractors.ms_modern.ooxml_shared import (
    OOXMLZipContext,
    get_image_content_type,
    get_image_pixel_dimensions,
)
from sharepoint2text.parsing.extractors.util.encryption import is_ooxml_encrypted
from sharepoint2text.parsing.extractors.util.zip_bomb import validate_zip_bytesio
from sharepoint2text.parsing.models import (
    CellValue,
    ContentUnit,
    DocumentMetadata,
    ExtractedDocument,
    ImageAsset,
    JsonValue,
    Table,
)

logger = logging.getLogger(__name__)

# Datetime types for isinstance check
_DATETIME_TYPES = (datetime.datetime, datetime.date, datetime.time)


def _is_xlsb_path(path: str | None) -> bool:
    return bool(path and path.lower().endswith(".xlsb"))


# =============================================================================
# Cell value handling
# =============================================================================


def _get_cell_value(cell_value: Any) -> Any:
    """Convert cell value to appropriate Python type (datetime -> ISO string)."""
    if cell_value is None:
        return None
    if isinstance(cell_value, _DATETIME_TYPES):
        return cell_value.isoformat()
    return cell_value


def _format_value_for_display(value: Any) -> str:
    """Format a value as string for text table display."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _is_cell_non_empty(val: Any) -> bool:
    """Check if a cell value is non-empty."""
    return val is not None and (not isinstance(val, str) or val.strip() != "")


def _is_meaningful_value(val: Any) -> bool:
    """Check if value is meaningful (not None, empty, or 'Unnamed:' placeholder)."""
    if val is None:
        return False
    if isinstance(val, str):
        return bool(val.strip()) and not val.startswith("Unnamed: ")
    return True


def _is_table_name_row(row: list[Any]) -> bool:
    """Check if row has exactly one meaningful cell (table name pattern)."""
    non_empty = sum(1 for val in row if _is_meaningful_value(val))
    return non_empty == 1 and len(row) > 1


# =============================================================================
# Sheet data extraction
# =============================================================================


def _read_sheet_data(ws: Worksheet) -> list[list[Any]]:
    """Read and trim sheet data in one pass over worksheet rows."""
    rows: list[list[Any]] = []
    last_data_row = 0
    last_data_col = 0

    for row in ws.iter_rows(values_only=True):
        normalized_row = list(row)
        row_last_data_col = 0
        for idx in range(len(normalized_row) - 1, -1, -1):
            if _is_cell_non_empty(normalized_row[idx]):
                row_last_data_col = idx + 1
                break

        rows.append(normalized_row)
        if row_last_data_col > 0:
            last_data_row = len(rows)
            if row_last_data_col > last_data_col:
                last_data_col = row_last_data_col

    if last_data_row == 0 or last_data_col == 0:
        return []

    trimmed_rows = [row[:last_data_col] for row in rows[:last_data_row]]
    headers = [
        (
            f"Unnamed: {i}"
            if val is None or (isinstance(val, str) and not val.strip())
            else str(val)
        )
        for i, val in enumerate(trimmed_rows[0])
    ]

    all_rows: list[list[Any]] = [headers]
    for row in trimmed_rows[1:]:
        all_rows.append([_get_cell_value(val) for val in row])

    return all_rows


def _format_sheet_as_text(all_rows: list[list[Any]]) -> str:
    """Format sheet data as an aligned text table."""
    if not all_rows:
        return ""

    num_cols = max(len(row) for row in all_rows)
    col_widths = [0] * num_cols

    # Format all values and calculate column widths in one pass
    formatted_rows: list[list[str]] = []
    for row in all_rows:
        formatted_row = [
            _format_value_for_display(row[i] if i < len(row) else None)
            for i in range(num_cols)
        ]
        for i, val in enumerate(formatted_row):
            if len(val) > col_widths[i]:
                col_widths[i] = len(val)
        formatted_rows.append(formatted_row)

    return "\n".join(
        " ".join(val.rjust(col_widths[i]) for i, val in enumerate(row))
        for row in formatted_rows
    )


# =============================================================================
# Metadata extraction
# =============================================================================


def _extract_metadata_from_workbook(wb: Any) -> DocumentMetadata:
    """Extract document metadata from an openpyxl workbook properties object."""
    props = wb.properties

    extra_properties: dict[str, JsonValue] = {}
    if props.lastModifiedBy:
        extra_properties["xlsx.last_modified_by"] = str(props.lastModifiedBy)
    if props.revision:
        extra_properties["xlsx.revision"] = str(props.revision)
    metadata = DocumentMetadata(
        title=props.title or None,
        subject=props.description or None,
        author=props.creator or None,
        created=(
            props.created.isoformat()
            if isinstance(props.created, datetime.datetime)
            else None
        ),
        modified=(
            props.modified.isoformat()
            if isinstance(props.modified, datetime.datetime)
            else None
        ),
        keywords=[
            item.strip()
            for item in str(props.keywords or "").replace(";", ",").split(",")
            if item.strip()
        ],
        language=props.language or None,
        properties=extra_properties,
    )
    return metadata


def _read_metadata(file_like: io.BytesIO) -> DocumentMetadata:
    """Extract document metadata from XLSX core properties."""
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)
    try:
        return _extract_metadata_from_workbook(wb)
    finally:
        wb.close()


# =============================================================================
# Image extraction
# =============================================================================


def _resolve_drawing_path(target: str) -> str:
    """Normalize drawing relationship targets to ZIP paths."""
    if target.startswith("/"):
        return target[1:]
    if target.startswith(".."):
        return "xl/" + target[3:]
    return "xl/worksheets/" + target


def _resolve_image_path(target: str) -> str:
    """Normalize image relationship targets to ZIP paths."""
    if target.startswith("/"):
        return target[1:]
    return "xl/media/" + target.rsplit("/", 1)[-1]


def _extract_images_from_zip(
    file_like: io.BytesIO, sheet_names: list[str]
) -> dict[int, list[ImageAsset]]:
    """Extract all images from XLSX by parsing the ZIP archive directly."""
    images_by_sheet: dict[int, list[ImageAsset]] = {}
    image_counter = 0

    ctx = OOXMLZipContext(file_like)
    try:
        # Build mapping of sheet index to drawing file
        sheet_to_drawing: dict[int, str] = {}
        for sheet_idx in range(len(sheet_names)):
            rels_path = f"xl/worksheets/_rels/sheet{sheet_idx + 1}.xml.rels"
            for rel in ctx.read_relationships_if_exists(rels_path):
                if "drawing" in rel["type"]:
                    sheet_to_drawing[sheet_idx] = _resolve_drawing_path(rel["target"])
                    break

        # Process each drawing file
        for sheet_idx, drawing_path in sheet_to_drawing.items():
            drawing_root = ctx.read_xml_root_if_exists(drawing_path)
            if drawing_root is None:
                continue

            # Parse drawing relationships to get image file paths
            drawing_rels_path = drawing_path.replace(
                "drawings/", "drawings/_rels/"
            ).replace(".xml", ".xml.rels")

            rid_to_image: dict[str, str] = {}
            for rel in ctx.read_relationships_if_exists(drawing_rels_path):
                if "image" in rel["type"]:
                    rid_to_image[rel["id"]] = _resolve_image_path(rel["target"])

            sheet_images: list[ImageAsset] = []

            for anchor_type in ANCHOR_TYPES:
                for anchor in drawing_root.iter(anchor_type):
                    pic = anchor.find(XDR_PIC)
                    if pic is None:
                        continue

                    try:
                        # Get dimensions from ext element
                        width, height = 0, 0
                        ext = anchor.find(XDR_EXT)
                        if ext is not None:
                            try:
                                width = int(ext.get("cx", "0")) // EMU_PER_PIXEL
                                height = int(ext.get("cy", "0")) // EMU_PER_PIXEL
                            except ValueError:
                                pass

                        # Get caption and description
                        caption, description = "", ""
                        nvPicPr = pic.find(XDR_NVPICPR)
                        if nvPicPr is not None:
                            cNvPr = nvPicPr.find(XDR_CNVPR)
                            if cNvPr is not None:
                                caption = cNvPr.get("name", "")
                                description = cNvPr.get("descr", "")

                        # Get the blip reference
                        blipFill = pic.find(XDR_BLIPFILL)
                        if blipFill is None:
                            continue

                        blip = blipFill.find(A_BLIP_XLSX)
                        if blip is None:
                            continue

                        embed_rid = blip.get(R_EMBED_XLSX, "")
                        if not embed_rid or embed_rid not in rid_to_image:
                            continue

                        image_path = rid_to_image[embed_rid]
                        image_bytes = ctx.read_bytes_if_exists(image_path)
                        if image_bytes is None:
                            continue
                        filename = image_path.rsplit("/", 1)[-1]

                        if width <= 0 or height <= 0:
                            dims = get_image_pixel_dimensions(image_bytes)
                            width = dims[0] or 0 if dims else 0
                            height = dims[1] or 0 if dims else 0

                        image_counter += 1
                        sheet_images.append(
                            ImageAsset(
                                number=image_counter,
                                filename=filename,
                                media_type=get_image_content_type(filename),
                                data=image_bytes,
                                width=width or None,
                                height=height or None,
                                caption=caption or None,
                                description=description or None,
                                properties={
                                    "xlsx.sheet_index": sheet_idx + 1,
                                    "xlsx.size_bytes": len(image_bytes),
                                },
                            )
                        )

                    except (KeyError, ValueError, OSError) as e:
                        logger.debug("Failed to extract image from drawing: %s", e)

            if sheet_images:
                images_by_sheet[sheet_idx] = sheet_images

    except (zipfile.BadZipFile, KeyError, ValueError, OSError) as e:
        logger.debug("Failed to extract images from XLSX: %s", e)
    finally:
        ctx.close()

    return images_by_sheet


# =============================================================================
# Content extraction
# =============================================================================


def _read_content(file_like: io.BytesIO) -> list[ContentUnit]:
    """Read all sheets from XLSX file and extract content."""
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        sheets = _read_content_from_workbook(wb, sheet_names)
    finally:
        wb.close()

    file_like.seek(0)
    images_by_sheet = _extract_images_from_zip(file_like, sheet_names)
    for sheet_idx, sheet_images in images_by_sheet.items():
        if sheet_idx < len(sheets):
            sheets[sheet_idx].images = sheet_images

    return sheets


def _read_content_from_workbook(wb: Any, sheet_names: list[str]) -> list[ContentUnit]:
    """Read all sheets from an openpyxl workbook and extract content."""
    sheets: list[ContentUnit] = []

    for sheet_number, sheet_name in enumerate(sheet_names, start=1):
        ws = wb[sheet_name]
        all_rows = _read_sheet_data(ws)
        text = _format_sheet_as_text(all_rows)

        # Skip first row if it's just a table name
        data_rows = (
            all_rows[1:] if all_rows and _is_table_name_row(all_rows[0]) else all_rows
        )

        sheets.append(
            ContentUnit(
                number=sheet_number,
                kind="sheet",
                title=str(sheet_name),
                text=str(sheet_name) + "\n" + text.strip(),
                tables=[
                    Table(
                        rows=cast(list[list[CellValue]], data_rows),
                        name=str(sheet_name),
                    )
                ],
            )
        )
    return sheets


# =============================================================================
# Main entry point
# =============================================================================


def _read_xlsb_sheet(workbook: Any, sheet_index: int, sheet_name: str) -> ContentUnit:
    """Extract cell values and display text from one XLSB worksheet."""
    rows: list[list[Any]] = []
    with workbook.get_sheet(sheet_index) as worksheet:
        for row in worksheet.rows(sparse=False):
            values = [_get_cell_value(getattr(cell, "v", None)) for cell in row]
            if any(_is_cell_non_empty(value) for value in values):
                rows.append(values)

    table_rows = cast(list[list[CellValue]], rows)
    return ContentUnit(
        number=sheet_index,
        kind="sheet",
        title=sheet_name,
        text=sheet_name + "\n" + _format_sheet_as_text(rows).strip(),
        tables=[Table(rows=table_rows, name=sheet_name)],
    )


def _read_xlsb(file_like: io.BytesIO, path: str | None = None) -> ExtractedDocument:
    """Extract row-accurate worksheet content from an XLSB workbook."""
    with tempfile.TemporaryDirectory() as temp_directory:
        workbook_path = Path(temp_directory) / "workbook.xlsb"
        file_like.seek(0)
        with workbook_path.open("wb") as workbook_file:
            shutil.copyfileobj(file_like, workbook_file)

        with open_workbook(str(workbook_path)) as workbook:
            sheet_names = [str(name) for name in workbook.sheets]
            sheets = [
                _read_xlsb_sheet(workbook, sheet_index, sheet_name)
                for sheet_index, sheet_name in enumerate(sheet_names, start=1)
            ]

    return ExtractedDocument(
        format="xlsb",
        source=source_metadata(path),
        units=sheets,
    )


def read_xlsx(
    file_like: io.BytesIO, path: str | None = None, *, ignore_images: bool = False
) -> Generator[ExtractedDocument, Any, None]:
    """
    Extract all relevant content from an Excel XLSX or XLSB file.

    Uses a generator pattern for API consistency. Excel files yield exactly one
    canonical document containing sheets, metadata, and images.

    Args:
        file_like: BytesIO object containing the XLSX or XLSB file data.
        path: Optional path to the source file for metadata.
        ignore_images: If True, skip image extraction.
    """
    try:
        file_like.seek(0)
        if is_ooxml_encrypted(file_like):
            raise ExtractionFileEncryptedError(
                "XLSX is encrypted or password-protected"
            )

        if _is_xlsb_path(path):
            file_like.seek(0)
            validate_zip_bytesio(file_like, source="read_xlsb")
            file_like.seek(0)
            content = _read_xlsb(file_like, path=path)
            logger.debug(
                "Extracted XLSB: sheets=%d, rows=%d",
                len(content.units),
                sum(len(unit.tables[0].rows) for unit in content.units if unit.tables),
            )
            yield content
            return

        file_like.seek(0)
        validate_zip_bytesio(file_like, source="read_xlsx")

        file_like.seek(0)
        wb = load_workbook(file_like, read_only=True, data_only=True)
        try:
            metadata = _extract_metadata_from_workbook(wb)
            sheet_names = list(wb.sheetnames)
            sheets = _read_content_from_workbook(wb, sheet_names)
        finally:
            wb.close()

        if not ignore_images:
            file_like.seek(0)
            images_by_sheet = _extract_images_from_zip(file_like, sheet_names)
            for sheet_idx, sheet_images in images_by_sheet.items():
                if sheet_idx < len(sheets):
                    sheets[sheet_idx].images = sheet_images

        total_rows = sum(len(unit.tables[0].rows) for unit in sheets if unit.tables)
        total_images = sum(len(sheet.images) for sheet in sheets)
        logger.debug(
            "Extracted XLSX: sheets=%d, rows=%d, images=%d",
            len(sheets),
            total_rows,
            total_images,
        )

        source_format = Path(path).suffix.lower().lstrip(".") if path else "xlsx"
        yield ExtractedDocument(
            format=source_format or "xlsx",
            source=source_metadata(path),
            metadata=metadata,
            units=sheets,
        )
    except ExtractionError:
        raise
    except (zipfile.BadZipFile, KeyError, ValueError, OSError) as exc:
        raise ExtractionFailedError("Failed to extract XLSX file", cause=exc) from exc
